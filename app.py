from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from flask_cors import CORS
import sqlite3, io, os, secrets
from urllib.parse import urlencode
from datetime import datetime
from collections import defaultdict
from flask_bcrypt import Bcrypt
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

app = Flask(__name__)

# Load environment variables
from dotenv import load_dotenv
load_dotenv()

app.secret_key = os.getenv("SECRET_KEY", "supersecretkey")
bcrypt = Bcrypt(app)

# Session cookies must survive the Google redirect round-trip
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_HTTPONLY"] = True

# Enable CORS for deployment
CORS(app, supports_credentials=True)

# Google OAuth Configuration
GOOGLE_CLIENT_ID = os.getenv("GOOGLE_CLIENT_ID", "")
GOOGLE_CLIENT_SECRET = os.getenv("GOOGLE_CLIENT_SECRET", "")


def get_google_oauth_redirect_uri():
    """
    Must match EXACTLY what is registered in Google Cloud Console (Authorized redirect URIs).

    Prefer OAUTH_REDIRECT_URI in .env when:
    - You use both localhost and 127.0.0.1 (pick one URI and always use that URL in the browser)
    - The app runs behind a reverse proxy and request.url_root is wrong → set full https callback URL

    Otherwise we build from the current request origin so the redirect matches the address bar.
    """
    explicit = (os.getenv("OAUTH_REDIRECT_URI") or "").strip()
    if explicit:
        return explicit.rstrip("/")
    base = (request.url_root or "").rstrip("/")
    path = url_for("oauth2callback")
    if not base:
        return url_for("oauth2callback", _external=True, _scheme=request.scheme).rstrip("/")
    if path.startswith("http"):
        return path.rstrip("/")
    return (base + path).rstrip("/")

# ---------------------- DB Setup ----------------------
def init_db():
    conn = sqlite3.connect("expenses.db")
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE,
                    password TEXT,
                    google_id TEXT,
                    picture TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS expenses (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER,
                    date TEXT,
                    name TEXT,
                    category TEXT,
                    amount REAL)""")
    c.execute("""CREATE TABLE IF NOT EXISTS goals (
                    user_id INTEGER PRIMARY KEY,
                    goal_amount REAL DEFAULT 0)""")
    # Migration: add columns that may not exist in older databases
    for migration in [
        "ALTER TABLE users ADD COLUMN google_id TEXT",
        "ALTER TABLE users ADD COLUMN picture TEXT",
    ]:
        try:
            c.execute(migration)
        except sqlite3.OperationalError:
            pass  # Column already exists — safe to ignore
    conn.commit()
    conn.close()

init_db()

def get_db():
    conn = sqlite3.connect("expenses.db")
    conn.row_factory = sqlite3.Row
    return conn

# ---------------------- Auth Routes ----------------------
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = bcrypt.generate_password_hash(request.form['password']).decode('utf-8')
        conn = get_db()
        try:
            conn.execute("INSERT INTO users (username, password) VALUES (?, ?)", (username, password))
            conn.commit()
            flash("Registration successful! Please login.", "success")
            return redirect(url_for('login'))
        except:
            flash("Username already exists.", "danger")
        finally:
            conn.close()
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
        conn.close()
        if user and bcrypt.check_password_hash(user['password'], password):
            user_dict = dict(user)
            session['user_id'] = user_dict['id']
            session['username'] = user_dict['username']
            session['picture'] = user_dict.get('picture')  # None for password-only accounts
            return redirect(url_for('index'))
        else:
            flash("Invalid username or password.", "danger")
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash("Logged out successfully.", "info")
    return redirect(url_for('login'))

# ---------------------- Google OAuth Routes ----------------------
@app.route('/google_login')
def google_login():
    if not GOOGLE_CLIENT_ID:
        flash("Google OAuth is not configured. Please set GOOGLE_CLIENT_ID in .env", "warning")
        return redirect(url_for('login'))
    if not GOOGLE_CLIENT_SECRET:
        flash("Google OAuth is not configured. Please set GOOGLE_CLIENT_SECRET in .env", "warning")
        return redirect(url_for('login'))

    state = secrets.token_urlsafe(32)
    session["oauth_state"] = state
    session.modified = True

    redirect_uri = get_google_oauth_redirect_uri()
    params = {
        "client_id": GOOGLE_CLIENT_ID,
        "redirect_uri": redirect_uri,
        "response_type": "code",
        "scope": "openid email profile",
        "state": state,
        "access_type": "offline",
        "prompt": "select_account",
    }
    google_auth_url = "https://accounts.google.com/o/oauth2/v2/auth?" + urlencode(params)
    return redirect(google_auth_url)

@app.route('/oauth2callback')
def oauth2callback():
    code = request.args.get('code')
    state = request.args.get('state')
    oauth_err = request.args.get("error")
    oauth_err_desc = request.args.get("error_description")

    if oauth_err:
        flash(
            f"Google sign-in was cancelled or denied: {oauth_err_desc or oauth_err}",
            "danger",
        )
        return redirect(url_for("login"))

    if not state or state != session.get('oauth_state'):
        flash("Invalid state parameter (try signing in again).", "danger")
        return redirect(url_for('login'))
    
    if not code:
        flash("Authorization failed (no code returned).", "danger")
        return redirect(url_for('login'))
    
    try:
        redirect_uri = get_google_oauth_redirect_uri()
        token_url = "https://oauth2.googleapis.com/token"
        data = {
            "client_id": GOOGLE_CLIENT_ID,
            "client_secret": GOOGLE_CLIENT_SECRET,
            "code": code,
            "grant_type": "authorization_code",
            "redirect_uri": redirect_uri,
        }
        
        token_response = requests.post(token_url, data=data, timeout=30)
        token_data = token_response.json()
        
        if "access_token" not in token_data:
            detail = token_data.get("error_description") or token_data.get("error") or token_response.text[:500]
            flash(f"Failed to obtain access token from Google: {detail}", "danger")
            return redirect(url_for('login'))
        
        access_token = token_data["access_token"]
        
        userinfo_url = "https://www.googleapis.com/oauth2/v2/userinfo"
        headers = {"Authorization": f"Bearer {access_token}"}
        userinfo_response = requests.get(userinfo_url, headers=headers, timeout=30)
        userinfo = userinfo_response.json()
        
        google_id = userinfo.get("id")
        email = userinfo.get("email")
        if not email:
            flash("Google did not return an email for this account. Check OAuth consent screen scopes.", "danger")
            return redirect(url_for("login"))
        picture = userinfo.get("picture")
        
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE google_id = ? OR username = ?", 
                           (google_id, email)).fetchone()
        
        if not user:
            password = bcrypt.generate_password_hash(os.urandom(24).hex()).decode('utf-8')
            conn.execute("INSERT INTO users (username, password, google_id, picture) VALUES (?, ?, ?, ?)",
                       (email, password, google_id, picture))
            conn.commit()
            user = conn.execute("SELECT * FROM users WHERE google_id = ?", (google_id,)).fetchone()
        
        conn.close()
        
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['picture'] = user['picture']
        
        flash(f"Welcome {user['username']}!", "success")
        return redirect(url_for('index'))
        
    except Exception as e:
        flash(f"Authentication error: {str(e)}", "danger")
        return redirect(url_for('login'))

@app.route('/coming-soon')
def coming_soon():
    return render_template('coming_soon.html')

# ---------------------- Dashboard with Filters ----------------------
@app.route('/', methods=['GET', 'POST'])
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db()

    filter_category = request.args.get('category', '').strip()
    filter_month = request.args.get('month', 'All')
    filter_year = request.args.get('year', 'All')
    filter_name = request.args.get('name', '').strip()

    query = "SELECT * FROM expenses WHERE user_id = ?"
    params = [session['user_id']]

    if filter_category and filter_category.lower() != "all":
        query += " AND LOWER(category) = ?"
        params.append(filter_category.lower())

    if filter_month != "All":
        query += " AND strftime('%m', date) = ?"
        params.append(f"{int(filter_month):02d}")

    if filter_year != "All":
        query += " AND strftime('%Y', date) = ?"
        params.append(filter_year)

    if filter_name:
        query += " AND LOWER(name) LIKE ?"
        params.append(f"%{filter_name.lower()}%")

    query += " ORDER BY date DESC"

    if not (filter_category or filter_month != "All" or filter_year != "All" or filter_name):
        query += " LIMIT 5"

    expenses = conn.execute(query, params).fetchall()

    total_income = sum(e['amount'] for e in conn.execute(
        "SELECT amount FROM expenses WHERE user_id = ? AND LOWER(category) = 'income'",
        (session['user_id'],)
    ))
    total_expense = sum(e['amount'] for e in conn.execute(
        "SELECT amount FROM expenses WHERE user_id = ? AND LOWER(category) != 'income'",
        (session['user_id'],)
    ))
    balance = total_income - total_expense

    g = conn.execute("SELECT goal_amount FROM goals WHERE user_id = ?", (session['user_id'],)).fetchone()
    goal = g['goal_amount'] if g else 0
    progress = (balance / goal * 100) if goal > 0 else 0

    years = [r['y'] for r in conn.execute(
        "SELECT DISTINCT strftime('%Y', date) as y FROM expenses WHERE user_id = ? ORDER BY y DESC",
        (session['user_id'],)
    ) if r['y']]

    conn.close()

    return render_template(
        'index.html',
        expenses=expenses,
        total_income=total_income,
        total_expense=total_expense,
        balance=balance,
        goal=goal,
        progress=progress,
        filter_category=filter_category,
        filter_month=filter_month,
        filter_year=filter_year,
        filter_name=filter_name,
        years=years
    )


# ---------------------- Add Expense or Income ----------------------
@app.route('/add', methods=['POST'])
def add():
    if 'user_id' not in session:
        print("[ADD EXPENSE] ❌ User not logged in")
        return jsonify({"error": "Not logged in", "success": False}), 401
    
    try:
        # Extract and validate form data
        date = request.form.get('date', '').strip()
        name = request.form.get('name', '').strip()
        category = request.form.get('category', '').strip()
        amount_str = request.form.get('amount', '').strip()
        
        # Validation checks
        if not date:
            print(f"[ADD EXPENSE] ❌ Missing date")
            return jsonify({"error": "Date is required", "success": False}), 400
        
        if not name:
            print(f"[ADD EXPENSE] ❌ Missing name")
            return jsonify({"error": "Description is required", "success": False}), 400
        
        if not category:
            print(f"[ADD EXPENSE] ❌ Missing category")
            return jsonify({"error": "Category is required", "success": False}), 400
        
        if not amount_str:
            print(f"[ADD EXPENSE] ❌ Missing amount")
            return jsonify({"error": "Amount is required", "success": False}), 400
        
        try:
            amount = float(amount_str)
            if amount <= 0:
                print(f"[ADD EXPENSE] ❌ Invalid amount: {amount}")
                return jsonify({"error": "Amount must be greater than 0", "success": False}), 400
        except ValueError:
            print(f"[ADD EXPENSE] ❌ Invalid amount format: {amount_str}")
            return jsonify({"error": "Amount must be a valid number", "success": False}), 400
        
        # Validate date format
        try:
            datetime.strptime(date, "%Y-%m-%d")
        except ValueError:
            print(f"[ADD EXPENSE] ❌ Invalid date format: {date}")
            return jsonify({"error": "Date must be in YYYY-MM-DD format", "success": False}), 400
        
        # Insert into database
        conn = get_db()
        c = conn.cursor()
        c.execute("INSERT INTO expenses (user_id, date, name, category, amount) VALUES (?, ?, ?, ?, ?)",
                  (session['user_id'], date, name, category, amount))
        conn.commit()
        
        # Get the inserted expense ID
        expense_id = c.lastrowid
        conn.close()
        
        print(f"[ADD EXPENSE] ✅ Success - User {session['user_id']}: {name} ({category}) ₹{amount:.2f} on {date}")
        return jsonify({
            "success": True,
            "message": f"✅ {name} added! ₹{amount:.2f}",
            "expense_id": expense_id
        }), 200
    
    except Exception as e:
        print(f"[ADD EXPENSE] ❌ Exception: {str(e)}")
        return jsonify({"error": f"Database error: {str(e)}", "success": False}), 400

@app.route('/add_income', methods=['POST'])
def add_income():
    if 'user_id' not in session:
        print("[ADD INCOME] ❌ User not logged in")
        return jsonify({"error": "Not logged in", "success": False}), 401
    
    try:
        # Extract and validate form data
        amount_str = request.form.get('income_amount', '').strip()
        
        if not amount_str:
            print("[ADD INCOME] ❌ Missing amount")
            return jsonify({"error": "Amount is required", "success": False}), 400
        
        try:
            amount = float(amount_str)
            if amount <= 0:
                print(f"[ADD INCOME] ❌ Invalid amount: {amount}")
                return jsonify({"error": "Amount must be greater than 0", "success": False}), 400
        except ValueError:
            print(f"[ADD INCOME] ❌ Invalid amount format: {amount_str}")
            return jsonify({"error": "Amount must be a valid number", "success": False}), 400
        
        today = datetime.now().strftime("%Y-%m-%d")
        
        conn = get_db()
        c = conn.cursor()
        c.execute("INSERT INTO expenses (user_id, date, name, category, amount) VALUES (?, ?, ?, ?, ?)",
                  (session['user_id'], today, "Income", "Income", amount))
        conn.commit()
        conn.close()
        
        print(f"[ADD INCOME] ✅ Success - User {session['user_id']}: ₹{amount:.2f} on {today}")
        return jsonify({
            "success": True,
            "message": f"✅ Income added! ₹{amount:.2f}",
        }), 200
    
    except Exception as e:
        print(f"[ADD INCOME] ❌ Exception: {str(e)}")
        return jsonify({"error": f"Database error: {str(e)}", "success": False}), 400

@app.route('/set_goal', methods=['POST'])
def set_goal():
    if 'user_id' not in session:
        print("[SET GOAL] ❌ User not logged in")
        return jsonify({"error": "Not logged in", "success": False}), 401
    
    try:
        goal_str = request.form.get('goal_amount', '').strip()
        
        if not goal_str:
            print("[SET GOAL] ❌ Missing goal amount")
            return jsonify({"error": "Goal amount is required", "success": False}), 400
        
        try:
            goal = float(goal_str)
            if goal < 0:
                print(f"[SET GOAL] ❌ Invalid goal: {goal}")
                return jsonify({"error": "Goal must be greater than or equal to 0", "success": False}), 400
        except ValueError:
            print(f"[SET GOAL] ❌ Invalid goal format: {goal_str}")
            return jsonify({"error": "Goal must be a valid number", "success": False}), 400
        
        conn = get_db()
        conn.execute("INSERT OR REPLACE INTO goals (user_id, goal_amount) VALUES (?, ?)", 
                    (session['user_id'], goal))
        conn.commit()
        conn.close()
        
        print(f"[SET GOAL] ✅ Success - User {session['user_id']}: Goal set to ₹{goal:.2f}")
        return jsonify({
            "success": True,
            "message": f"✅ Savings goal set to ₹{goal:.2f}",
        }), 200
    
    except Exception as e:
        print(f"[SET GOAL] ❌ Exception: {str(e)}")
        return jsonify({"error": f"Database error: {str(e)}", "success": False}), 400

@app.route('/get_expenses')
def get_expenses():
    if 'user_id' not in session:
        return jsonify({"error": "Not logged in"}), 401
    
    conn = get_db()
    
    filter_category = request.args.get('category', '').strip()
    filter_month = request.args.get('month', 'All')
    filter_year = request.args.get('year', 'All')
    filter_name = request.args.get('name', '').strip()
    
    query = "SELECT * FROM expenses WHERE user_id = ?"
    params = [session['user_id']]
    
    if filter_category and filter_category.lower() != "all":
        query += " AND LOWER(category) = ?"
        params.append(filter_category.lower())
    
    if filter_month != "All":
        query += " AND strftime('%m', date) = ?"
        params.append(f"{int(filter_month):02d}")
    
    if filter_year != "All":
        query += " AND strftime('%Y', date) = ?"
        params.append(filter_year)
    
    if filter_name:
        query += " AND LOWER(name) LIKE ?"
        params.append(f"%{filter_name.lower()}%")
    
    query += " ORDER BY date DESC"
    
    if not (filter_category or filter_month != "All" or filter_year != "All" or filter_name):
        query += " LIMIT 5"
    
    expenses = conn.execute(query, params).fetchall()
    
    total_income = sum(e['amount'] for e in conn.execute(
        "SELECT amount FROM expenses WHERE user_id = ? AND LOWER(category) = 'income'",
        (session['user_id'],)
    ))
    total_expense = sum(e['amount'] for e in conn.execute(
        "SELECT amount FROM expenses WHERE user_id = ? AND LOWER(category) != 'income'",
        (session['user_id'],)
    ))
    balance = total_income - total_expense
    
    g = conn.execute("SELECT goal_amount FROM goals WHERE user_id = ?", (session['user_id'],)).fetchone()
    goal = g['goal_amount'] if g else 0
    progress = (balance / goal * 100) if goal > 0 else 0
    
    conn.close()
    
    return jsonify({
        "expenses": [dict(e) for e in expenses],
        "total_income": total_income,
        "total_expense": total_expense,
        "balance": balance,
        "goal": goal,
        "progress": progress
    }), 200

# ---------------------- Delete Expense ----------------------
@app.route('/delete/<int:id>', methods=['GET', 'POST'])
def delete(id):
    if 'user_id' not in session:
        print(f"[DELETE] ❌ User not logged in")
        return jsonify({"error": "Not logged in", "success": False}), 401
    
    try:
        conn = get_db()
        c = conn.cursor()
        
        # Verify the expense belongs to the current user
        expense = c.execute("SELECT user_id FROM expenses WHERE id = ?", (id,)).fetchone()
        if not expense:
            print(f"[DELETE] ❌ Expense {id} not found")
            return jsonify({"error": "Expense not found", "success": False}), 404
        
        if expense['user_id'] != session['user_id']:
            print(f"[DELETE] ❌ Unauthorized - User {session['user_id']} trying to delete expense {id}")
            return jsonify({"error": "Unauthorized", "success": False}), 403
        
        c.execute("DELETE FROM expenses WHERE id = ?", (id,))
        conn.commit()
        conn.close()
        
        print(f"[DELETE] ✅ Success - Expense {id} deleted by user {session['user_id']}")
        return jsonify({"success": True, "message": "✅ Expense deleted successfully!"}), 200
    
    except Exception as e:
        print(f"[DELETE] ❌ Exception: {str(e)}")
        return jsonify({"error": f"Database error: {str(e)}", "success": False}), 400

# ---------------------- Export as Excel ----------------------
@app.route('/export')
def export_data():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db()
    expenses = conn.execute("SELECT date, name, category, amount FROM expenses WHERE user_id = ? ORDER BY date DESC", 
                          (session['user_id'],)).fetchall()
    conn.close()

    df = pd.DataFrame(expenses, columns=['Date', 'Name', 'Category', 'Amount'])
    
    # Create Excel workbook with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.border = thin_border
            
            if r_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            else:
                if c_idx == 4:
                    cell.number_format = '₹#,##0.00'
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return send_file(
        excel_buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'expenses_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

# ---------------------- Comprehensive Stats Page ----------------------
@app.route('/stats')
def stats():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    conn = get_db()
    
    # Category breakdown
    categories = conn.execute(
        "SELECT category, SUM(amount) as total FROM expenses WHERE user_id = ? AND LOWER(category) != 'income' GROUP BY category ORDER BY total DESC",
        (session['user_id'],)
    ).fetchall()
    
    # Monthly trend
    monthly_data = conn.execute(
        "SELECT strftime('%Y-%m', date) as month, SUM(amount) as total "
        "FROM expenses WHERE user_id = ? AND LOWER(category) != 'income' GROUP BY month ORDER BY month ASC",
        (session['user_id'],)
    ).fetchall()
    
    # Overall statistics
    total_income = sum(e['amount'] for e in conn.execute(
        "SELECT amount FROM expenses WHERE user_id = ? AND LOWER(category) = 'income'",
        (session['user_id'],)
    ))
    total_expense = sum(e['amount'] for e in conn.execute(
        "SELECT amount FROM expenses WHERE user_id = ? AND LOWER(category) != 'income'",
        (session['user_id'],)
    ))
    balance = total_income - total_expense
    
    # Get top 5 expenses
    top_expenses = conn.execute(
        "SELECT name, category, amount, date FROM expenses WHERE user_id = ? AND LOWER(category) != 'income' ORDER BY amount DESC LIMIT 5",
        (session['user_id'],)
    ).fetchall()
    
    # Calculate average expense
    all_expenses = conn.execute(
        "SELECT amount FROM expenses WHERE user_id = ? AND LOWER(category) != 'income'",
        (session['user_id'],)
    ).fetchall()
    avg_expense = (total_expense / len(all_expenses)) if all_expenses else 0
    
    conn.close()
    
    return render_template(
        'stats.html',
        categories=categories,
        monthly_data=monthly_data,
        total_income=total_income,
        total_expense=total_expense,
        balance=balance,
        avg_expense=avg_expense,
        top_expenses=top_expenses
    )


# --------------------------
# Run App
# --------------------------
if __name__ == "__main__":
    app.run(debug=True)