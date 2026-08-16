import io
import sqlite3

from openpyxl import load_workbook

from app import app


def set_session(client, user_id=1):
    with client.session_transaction() as session:
        session['user_id'] = user_id
        session['username'] = 'tester'


def test_export_contains_date_day_time_and_amount_columns():
    conn = sqlite3.connect('expenses.db')
    conn.execute(
        "INSERT INTO expenses (user_id, date, time, name, category, amount) VALUES (?, ?, ?, ?, ?, ?)",
        (1, '2026-08-16', '13:12', 'll', 'Education', 99.0),
    )
    conn.execute(
        "INSERT INTO expenses (user_id, date, time, name, category, amount) VALUES (?, ?, ?, ?, ?, ?)",
        (1, '2026-07-24', '10:45', 'clothes', 'Miscellaneous', 840.0),
    )
    conn.commit()
    conn.close()

    client = app.test_client()
    set_session(client)

    response = client.get('/export')
    assert response.status_code == 200
    assert response.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    workbook_bytes = io.BytesIO(response.data)
    workbook = load_workbook(workbook_bytes)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]
    assert headers[:6] == ['Date', 'Day', 'Time', 'Name', 'Category', 'Amount']

    records = []
    for row in sheet.iter_rows(min_row=2, values_only=False):
        records.append([
            row[0].value,
            row[1].value,
            row[2].value,
            row[3].value,
            row[4].value,
            row[5].value,
        ])

    found = False
    for row_data in records:
        export_date = row_data[0]
        if hasattr(export_date, 'date'):
            date_value = export_date.date().isoformat()
        else:
            date_value = str(export_date).split('T')[0]

        if date_value == '2026-08-16' and row_data[1] == 'Sunday' and row_data[2].strftime('%H:%M') == '13:12' and row_data[3] == 'll' and row_data[4] == 'Education' and row_data[5] == 99.0:
            found = True
            break

    assert found is True
    assert sheet.max_column == 6
